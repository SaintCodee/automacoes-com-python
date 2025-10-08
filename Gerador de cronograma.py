from datetime import date
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import time
from datetime import datetime
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.edge.service import Service
import win32com.client as win32
from pathlib import Path
import os
import shutil  
import ttkbootstrap as ttk
from tkinter import messagebox, StringVar
from ttkbootstrap.constants import *
from ttkbootstrap.widgets import DateEntry
import json
# import config # Importa o arquivo de configurações sensíveis


# --- CONFIGURAÇÕES GLOBAIS E ANONIMIZAÇÃO ---

# Variável de contagem (mantida)
i = 0

# Caminhos Anonimizados e Universais
DADOS_SALVOS_PATH = 'dados_entrada.json'
CRONOGRAMA_MODELO = 'cronograma_modelo.xlsx'
OUTPUT_DIR = config.OUTPUT_DIR_LOCAL # Puxado do arquivo config.py

# Caminho seguro e universal para Downloads (para arquivos temporários)
DOWNLOADS_PATH_TEMP = os.path.join(os.path.expanduser("~"), "Downloads")
FILE_NAME_TEMP = 'ordens_servico.xls'
FILE_PATH_TEMP = os.path.join(DOWNLOADS_PATH_TEMP, FILE_NAME_TEMP)
FINAL_XLSX_NAME = 'ordens_servico_dados_equipamentos.xlsx'
TEMP_XLSX_PATH = os.path.join(DOWNLOADS_PATH_TEMP, FINAL_XLSX_NAME)
MOVE_TO_PATH = os.path.join(OUTPUT_DIR, FINAL_XLSX_NAME)

# --- Estilo das células (mantido) ---
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
alignment = Alignment(horizontal='center', vertical='center')
font = Font(name='Times New Roman', size=9)
fundo_verde = PatternFill(start_color="0000FF00", end_color="0000FF00", fill_type="solid")
fundo_amarelo = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")
negrito = Font(name='Times New Roman', size=9, bold=True)
font_data = Font(name='Calibri', size=20, bold=True)


# --- Funções auxiliares (mantidas) ---
def carregar_clientes_excel(caminho_arquivo='clientes.xlsx'):
    df = pd.read_excel(caminho_arquivo)
    return df['Cliente'].dropna().astype(str).tolist()

# --- Coleta de dados via GUI (mantida, mas com caminhos corrigidos) ---
def obter_dados():
    root = ttk.Window(themename='superhero')
    root.title("Dados de Entrada")
    root.geometry('600x700')
    root.minsize(500,600)
    root.resizable(False, False)

    # ... (Restante do código da GUI mantido) ...
    cliente_var = StringVar()
    setor_var = StringVar()
    validade_var = StringVar()
    tecnicos_var = StringVar()
    periodicidade_var = StringVar()
    plano_var = StringVar()
    resultados = {}

    # Carrega anteriores
    try:
        with open(DADOS_SALVOS_PATH, 'r', encoding='utf-8') as f:
            dados = json.load(f)
            cliente_var.set(dados.get('cliente',''))
            setor_var.set(dados.get('setor',''))
            validade_var.set(dados.get('validade',''))
            tecnicos_var.set(dados.get('tecnicos',''))
            periodicidade_var.set(dados.get('periodicidade',''))
            plano_var.set(dados.get('plano',''))
    except FileNotFoundError:
        pass

    lista_clientes = carregar_clientes_excel()
    def on_cliente_keyrelease(event):
        val = cliente_var.get().lower()
        cliente_combo['values'] = lista_clientes if not val else [c for c in lista_clientes if val in c.lower()]

    def confirmar():
        ci = data_inicial.entry.get()
        cf = data_final.entry.get()
        if not all([cliente_var.get(), setor_var.get(), validade_var.get(),
                     tecnicos_var.get(), periodicidade_var.get(), ci, cf]):
            messagebox.showerror("Erro","Preencha todos os campos obrigatórios!")
            return
        to_save = {
            'cliente': cliente_var.get(),
            'setor': setor_var.get(),
            'validade': validade_var.get(),
            'tecnicos': tecnicos_var.get(),
            'periodicidade': periodicidade_var.get(),
            'plano': plano_var.get(),
            'data_inicial': ci,
            'data_final': cf
        }
        with open(DADOS_SALVOS_PATH,'w',encoding='utf-8') as f:
            json.dump(to_save,f,ensure_ascii=False,indent=4)
        resultados.update(to_save)
        root.destroy()

    # Layout
    frame = ttk.Frame(root, padding=30)
    frame.pack(fill='both',expand=True)

    ttk.Label(frame,text="Cliente:").pack(anchor='w',padx=5)
    cliente_combo = ttk.Combobox(frame,textvariable=cliente_var,values=lista_clientes,width=40)
    cliente_combo.pack(fill='x',padx=5,pady=(0,10))
    cliente_combo.bind('<KeyRelease>',on_cliente_keyrelease)

    ttk.Label(frame,text="Setor:").pack(anchor='w',padx=5)
    ttk.Entry(frame,textvariable=setor_var,width=40).pack(fill='x',padx=5,pady=(0,10))

    ttk.Label(frame,text="Validade do Relatório:").pack(anchor='w',padx=5)
    ttk.Entry(frame,textvariable=validade_var,width=40).pack(fill='x',padx=5,pady=(0,10))

    ttk.Label(frame,text="Técnicos:").pack(anchor='w',padx=5)
    ttk.Entry(frame,textvariable=tecnicos_var,width=40).pack(fill='x',padx=5,pady=(0,10))

    ttk.Label(frame,text="Periodicidade:").pack(anchor='w',padx=5)
    ttk.Combobox(
        frame,textvariable=periodicidade_var,
        values=["Anual","Semestral","Trimestral","Mensal"],width=40
    ).pack(fill='x',padx=5,pady=(0,10))

    ttk.Label(frame,text="Plano (opcional):").pack(anchor='w',padx=5)
    ttk.Entry(frame,textvariable=plano_var,width=40).pack(fill='x',padx=5,pady=(0,10))

    ttk.Label(frame,text="Data Inicial:").pack(anchor='w',padx=5)
    data_inicial = DateEntry(frame,dateformat='%d/%m/%Y',firstweekday=6,width=12,bootstyle='success')
    data_inicial.pack(fill='x',padx=5,pady=(0,10))
    if setor_var.get(): pass
    if os.path.exists(DADOS_SALVOS_PATH):
        data_inicial.entry.delete(0,'end')
        with open(DADOS_SALVOS_PATH,'r') as f:di=json.load(f)['data_inicial'];data_inicial.entry.insert(0,di)

    ttk.Label(frame,text="Data Final:").pack(anchor='w',padx=5)
    data_final = DateEntry(frame,dateformat='%d/%m/%Y',firstweekday=6,width=12,bootstyle='success')
    data_final.pack(fill='x',padx=5,pady=(0,10))
    if os.path.exists(DADOS_SALVOS_PATH):
        data_final.entry.delete(0,'end')
        with open(DADOS_SALVOS_PATH,'r') as f:df=json.load(f)['data_final'];data_final.entry.insert(0,df)

    ttk.Button(frame,text="Gerar Cronograma",command=confirmar,bootstyle="success").pack(fill='x',padx=5,pady=20)

    root.mainloop()

    if not resultados:
        return None

    di = datetime.strptime(resultados['data_inicial'],'%d/%m/%Y').date()
    df = datetime.strptime(resultados['data_final'],'%d/%m/%Y').date()
    return (
        resultados['cliente'],resultados['setor'],resultados['validade'],
        resultados['tecnicos'],resultados['periodicidade'],di,df,resultados['plano']
    )


# --- Fluxo principal ---
if __name__ == '__main__':
    dados = obter_dados()
    if not dados:
        print("\n❌ Operação cancelada.")
    else:
        cliente,setor,validade,tecnicos,periodicidade,data_inicial,data_final,plano = dados
        print(f"\n✅ Gerando cronograma para {cliente} ({setor}) de {data_inicial} até {data_final}...")
        
        # Inicia a automação com URL anonimizada
        navegador = f.conexão()
        action = ActionChains(navegador)
        site = config.ARKMEDS_URL_EXPORT # URL anonimizada
        # f.login(navegador,site)
        
        # --- Preenchimento no site (usando WebDriverWait para robustez) ---
        
        # Abre o dropdown de Empresas
        WebDriverWait(navegador, 20).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="div_id_empresa"]/div/span/div/button'))
        ).click()
        
        # Digita o nome do cliente no campo de busca do dropdown
        campo_busca = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="div_id_empresa"]/div/span/div/ul/li[1]/div/input'))
        )
        campo_busca.clear()
        campo_busca.send_keys(cliente)
        time.sleep(1) # Pequeno delay para filtro
        
        # Clica na opção do cliente
        opcao_cliente = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//label[contains(., '" + cliente + "')]"))
        )
        opcao_cliente.click()
        action.send_keys(Keys.TAB).perform()

        # Tipo de Serviço (usando WebDriverWait)
        WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="div_id_tipo_servico"]/div/span/div/button/b'))
        ).click()
        time.sleep(1)
        navegador.find_element(By.XPATH, '//*[@id="div_id_tipo_servico"]/div/span/div/ul/li[1]/div/input').click()
        navegador.find_element(By.XPATH, '//*[@id="div_id_tipo_servico"]/div/span/div/ul/li[21]/a/label').click()  # Preventiva
        navegador.find_element(By.XPATH, '//*[@id="div_id_tipo_servico"]/div/span/div/ul/li[20]/a/label').click()  # Calibração
        action.send_keys(Keys.TAB).perform()

        # Plano (usando WebDriverWait)
        if plano != "":
            WebDriverWait(navegador, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="div_id_plano"]/div/span/div/button'))
            ).click()
            navegador.find_element(By.XPATH, '//*[@id="div_id_plano"]/div/span/div/ul/li[1]/div/input').click()
            action.send_keys(plano).perform()
            navegador.find_element(By.XPATH, '//*[@id="div_id_plano"]/div/span/div/ul/li[2]/a/label').click()
            action.send_keys(Keys.TAB).perform()

        # Estado da OS (usando WebDriverWait)
        WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="div_id_estado"]/div/span/div/button/b'))
        ).click()
        navegador.find_element(By.XPATH, '//*[@id="div_id_estado"]/div/span/div/ul/li[3]/a/label').click()  # Fechada
        action.send_keys(Keys.TAB).perform()

        # Intervalo de datas
        campo_intervalo = WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="id_intervalo"]'))
        )
        campo_intervalo.clear()
        campo_intervalo.click()
        intervalo_datas = data_inicial.strftime('%d/%m/%Y') + ' - ' + data_final.strftime('%d/%m/%Y')
        action.send_keys(intervalo_datas).perform()
        action.send_keys(Keys.TAB).perform()

        # Colunas do relatório (Usando WebDriverWait)
        WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="div_id_colunas_opcionais"]/div/span/div/button'))
        ).click()
        
        # Lista de XPATHS para as colunas (melhorando a legibilidade)
        xpaths_colunas = [
            '//*[@id="div_id_colunas_opcionais"]/div/span/div/ul/li[9]/a/label', # Tipo Eq
            '//*[@id="div_id_colunas_opcionais"]/div/span/div/ul/li[10]/a/label', # NS
            '//*[@id="div_id_colunas_opcionais"]/div/span/div/ul/li[12]/a/label', # Patrimonio
            '//*[@id="div_id_colunas_opcionais"]/div/span/div/ul/li[13]/a/label', # Modelo
            '//*[@id="div_id_colunas_opcionais"]/div/span/div/ul/li[14]/a/label'  # Fabricante
        ]
        
        for xpath in xpaths_colunas:
            WebDriverWait(navegador, 5).until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
            
        action.send_keys(Keys.TAB).perform()
        WebDriverWait(navegador, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="main-content-wrapper"]/div/div[2]/form/div[2]/button/strong'))
        ).click()
        time.sleep(10) # Tempo de espera para o download

        # --- Processamento de Arquivos (Caminhos Anonimizados) ---
        
        # Leitura e conversão
        data_xls = pd.read_excel(FILE_PATH_TEMP, index_col=None)
        base_name = os.path.splitext(FILE_PATH_TEMP)[0]
        data_xls.to_excel(TEMP_XLSX_PATH, index=False)
        
        # Limpeza e movimentação (usando caminhos anominizados)
        os.remove(FILE_PATH_TEMP)
        
        # Cria a pasta de saída se ela não existir
        os.makedirs(os.path.dirname(MOVE_TO_PATH), exist_ok=True)
        shutil.move(TEMP_XLSX_PATH, MOVE_TO_PATH)
        
        # --- Manipulação do Excel (OpenPyXL) ---
        
        mes = data_inicial.month - 1
        wb = load_workbook(CRONOGRAMA_MODELO)
        wa = load_workbook(MOVE_TO_PATH)
        modelo = wb['Cronograma']
        equipamentos = wa['Sheet1']

        # Preenchimento do cabeçalho
        modelo['B6'] = setor
        modelo['D6'] = cliente
        modelo['I6'] = validade
        modelo['B7'] = tecnicos
        equipamentos.delete_rows(1)
        Tipo_eq = equipamentos['C']
        Serial = equipamentos['D']
        Patrimonio = equipamentos['E']
        Modelo = equipamentos['F']
        Fabricante = equipamentos['G']
        ws = wb.active

        # Lógica de mesclagem e preenchimento de meses (mantida)
        ws.merge_cells(start_row=7, start_column=7, end_row=8, end_column=19-mes-1)
        ws.merge_cells(start_row=7, start_column=20-mes-1, end_row=8, end_column=19)
        ws.cell(row=7, column=20-mes-1, value=str(data_final.year)) # Ano do segundo ano (melhorado)
        
        for row in ws['G7:S8']:
            for cell in cell:
                cell.font = font_data
                cell.alignment = alignment
                
        meses = ['JAN','FEV','MAR','ABR','MAI','JUN','JUL','AGO','SET','OUT','NOV','DEZ',
                 'JAN','FEV','MAR','ABR','MAI','JUN','JUL','AGO','SET','OUT','NOV','DEZ']        
        
        i = 0
        for row in modelo['G9':'S9']:
            for cell in row:
                ws.cell(row=9, column=7 + i, value=meses[mes+i])
                i += 1
                
        ultimo_dado = equipamentos.max_row
        # ... (Restante do código de inserção de dados e formatação mantido) ...
        
        ## Inserção de dados da planilha de equipamentos para o cronograma ##
        i = 0
        for celula in Tipo_eq:
            modelo['A'+str(i+10)] = celula.value
            i += 1
        i = 0
        for celula in Serial:
            modelo['B'+str(i+10)] = celula.value
            i += 1
        i = 0
        for celula in Patrimonio:
            modelo['C'+str(i+10)] = celula.value
            i += 1
        i = 0
        for celula in Modelo:
            modelo['D'+str(i+10)] = celula.value
            i += 1
        i = 0
        for celula in Fabricante:
            modelo['E'+str(i+10)] = celula.value
            modelo['F'+str(i+10)] = periodicidade
            modelo['G'+str(i+10)] = "R"
            i += 1

        for row in modelo['A10':'S'+str(ultimo_dado+9)]:
            for cell in cell:
                cell.border = thin_border
                cell.alignment = alignment
                cell.font = font

        for row in modelo['G10':'G'+str(ultimo_dado+9)]:
            for cell in cell:
                cell.fill = fundo_verde
                cell.font = negrito
                
        ## Colocando os P's nos meses conforme a periodicidade ##
        if periodicidade == 'Anual':
            for row in modelo['S10':'S'+str(ultimo_dado+9)]:
                for cell in row:
                    cell.value = "P"
                    cell.fill = fundo_amarelo
                    cell.font = negrito

        if periodicidade == 'Semestral':
            for row in modelo['S10':'S'+str(ultimo_dado+9)]:
                for cell in row:
                    cell.value = "P"
                    cell.fill = fundo_amarelo
                    cell.font = negrito
            for row in modelo['M10':'M'+str(ultimo_dado+9)]:
                for cell in row:
                    cell.value = "P"
                    cell.fill = fundo_amarelo
                    cell.font = negrito
                    
        if periodicidade == 'Trimestral':
            for row in modelo['S10':'S'+str(ultimo_dado+9)]:
                for cell in row:
                    cell.value = "P"
                    cell.fill = fundo_amarelo
                    cell.font = negrito
            for row in modelo['M10':'M'+str(ultimo_dado+9)]:
                for cell in row:
                    cell.value = "P"
                    cell.fill = fundo_amarelo
                    cell.font = negrito
            for row in modelo['J10':'J'+str(ultimo_dado+9)]:
                for cell in row:
                    cell.value = "P"
                    cell.fill = fundo_amarelo
                    cell.font = negrito
            for row in modelo['P10':'P'+str(ultimo_dado+9)]:
                for cell in row:
                    cell.value = "P"
                    cell.fill = fundo_amarelo
                    cell.font = negrito
                    
        if periodicidade == "Mensal":
            for row in modelo['H10':'S'+str(ultimo_dado+9)]:
                for cell in row:
                    cell.value = "P"
                    cell.fill = fundo_amarelo
                    cell.font = negrito

        # --- Remoção de linhas duplicadas na planilha "Cronograma" (mantida) ---

        def remove_duplicate_rows(sheet, start_row, start_col, end_col):
            seen = set()
            rows_to_delete = []
            for row in range(start_row, sheet.max_row + 1):
                values = tuple(sheet.cell(row=row, column=col).value for col in range(start_col, end_col + 1))
                if values in seen:
                    rows_to_delete.append(row)
                else:
                    seen.add(values)
            # Delete from bottom to top to avoid shifting rows
            for row in reversed(rows_to_delete):
                sheet.delete_rows(row)

        remove_duplicate_rows(modelo, start_row=10, start_col=1, end_col=19)
                
        # Ordena alfabeticamente pelo Tipo de Equipamento (coluna A)
        dados = []
        for row in modelo.iter_rows(min_row=10, max_row=modelo.max_row, min_col=1, max_col=19, values_only=True):
            if any(row):
                dados.append(row)

        dados_ordenados = sorted(dados, key=lambda x: str(x[0]).upper() if x[0] is not None else "")

        for i in range(modelo.max_row, 9, -1):
            modelo.delete_rows(i)

        for idx, row in enumerate(dados_ordenados, start=10):
            for col, value in enumerate(row, start=1):
                modelo.cell(row=idx, column=col, value=value)

        # Agora APLIQUE A FORMATAÇÃO nas linhas já ordenadas
        for row in modelo['A10':'S'+str(modelo.max_row)]:
            for cell in cell:
                cell.border = thin_border
                cell.alignment = alignment
                cell.font = font

        # Coluna G (verde)
        for row in modelo['G10':'G'+str(modelo.max_row)]:
            for cell in cell:
                cell.fill = fundo_verde
                cell.font = negrito

        # Todas as células "P" nas colunas de meses (J até S) ficam amarelas
        for row in modelo.iter_rows(min_row=10, max_row=modelo.max_row, min_col=10, max_col=19):
            for cell in row:
                if cell.value == "P":
                    cell.fill = fundo_amarelo
                    cell.font = negrito

        # Salva o arquivo Excel com o nome do cliente e setor (caminho anonimizado)
        final_output_path = os.path.join(OUTPUT_DIR, cliente + ' - ' + setor + '.xlsx')
        wb.save(final_output_path)
        
        print(f"\n✅ Cronograma salvo em: {final_output_path}")
        input('Pressione ENTER para finalizar...')